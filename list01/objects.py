import os
import shutil
import openpyxl
from openpyxl.utils import range_boundaries

print(os.path.dirname(__file__))
LIST01_STATIC_FILES_PATH = os.path.join(os.path.dirname(__file__), 'list01-staticfiles')
SINGLE_TEMPLATE_PATH = os.path.join(LIST01_STATIC_FILES_PATH, 'List01_Template.xlsx')
LIST01_TEMPLATE = os.path.join(LIST01_STATIC_FILES_PATH, 'List01_Template_Output.xlsx')

class MakeupList01:
    shutil.copy2(SINGLE_TEMPLATE_PATH, LIST01_TEMPLATE)
    def __init__(self, single_template_path):
        self.single_template_path = single_template_path
        self.wb = openpyxl.load_workbook(single_template_path)
        self.ws = self.wb.active

    def make_up(self):
        copy_range = 'A93:W179'
        copy_range_start = 'A93'
        copy_range_end = 'W179'
        paste_range_start = 'W180'
        paste_start_cell = 'A180'

        for row_num, row in enumerate(self.ws[copy_range_start:copy_range_end], start=180):
            for col_num, cell in enumerate(row, start=1):
                self.ws.cell(row=row_num, column=col_num, value=cell.value)


        self.wb.save(LIST01_TEMPLATE)