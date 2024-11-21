import os
import sys
import warnings
import openpyxl
import shutil
import pandas as pd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
KAM12FILEREADER_DIR = os.path.join(PARENT_DIR, 'kam12filereader')

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))
sys.path.append(KAM12FILEREADER_DIR)

BASEACTIONS_DIR = os.path.dirname(SCRIPT_DIR)
PROJECT_DIR = os.path.dirname(BASEACTIONS_DIR)
STATICFILES_DIR = os.path.join(PROJECT_DIR, 'staticfiles')
EXCEL_DIR = os.path.join(STATICFILES_DIR, 'excel')

CSV_DIR = os.path.join(EXCEL_DIR, 'csv')
XLSX_DIR = os.path.join(EXCEL_DIR, 'xlsx')
ACC_REPORT_PATH = os.path.join(XLSX_DIR, 'acc_report.xlsx')
OUT_ACC_REPORT_PATH = os.path.join(XLSX_DIR, 'out_acc_report.xlsx')

SAD_FILE_PATH = os.path.join(XLSX_DIR, 'SAD.xlsx')
BUDGET_FILE_PATH = os.path.join(XLSX_DIR, 'Budget.xlsx')

from baseoperations.kam12filereader.main_operation import BudgetAccount, SADDetail

SAD = SADDetail(SAD_FILE_PATH)
BUDGET = BudgetAccount(BUDGET_FILE_PATH)


class AccReport:
    shutil.copy2(ACC_REPORT_PATH, OUT_ACC_REPORT_PATH)
    def __init__(self):
        self.acc_wb = openpyxl.load_workbook(OUT_ACC_REPORT_PATH)
        self.acc_ws = self.acc_wb.active


    def get_import_stat(self):     
        for i in range(8, 60, 1):
            self.acc_wb['1'].cell(row=i, column=5).value = SAD.get_value_by_group_id(value_code='khr_value', group_id=i-7, transaction='IM')
            self.acc_wb['1'].cell(row=i, column=6).value = SAD.get_value_by_group_id(value_code='cop', group_id=i-7, transaction='IM') \
                                                        + SAD.get_value_by_group_id(value_code='cpp', group_id=i-7, transaction='IM')
            self.acc_wb['1'].cell(row=i, column=7).value = SAD.get_value_by_group_id(value_code='atp', group_id=i-7, transaction='IM')
            self.acc_wb['1'].cell(row=i, column=8).value = SAD.get_value_by_group_id(value_code='sop', group_id=i-7, transaction='IM') \
                                                        + SAD.get_value_by_group_id(value_code='spp', group_id=i-7, transaction='IM')
            self.acc_wb['1'].cell(row=i, column=9).value = SAD.get_value_by_group_id(value_code='vop', group_id=i-7, transaction='IM') \
                                                        + SAD.get_value_by_group_id(value_code='vpp', group_id=i-7, transaction='IM') 
        self.acc_wb.save(OUT_ACC_REPORT_PATH)  
        self.acc_wb.close()  

    def acc_report_to_workbook(self):
        return ACC_REPORT_PATH                                        
    
    def export_stat(self):
        for i in range(9, 38, 1):
            self.acc_wb['3'].cell(row=i, column=5).value = SAD.get_value_by_group_id(value_code='khr_value', group_id=i-8, transaction='EX')
            self.acc_wb['3'].cell(row=i, column=8).value = SAD.get_value_by_group_id(value_code='khr_value', group_id=i-8, transaction='EX')


    def save_report(self):
        self.acc_wb.save(OUT_ACC_REPORT_PATH)