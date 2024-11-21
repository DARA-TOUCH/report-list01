import pandas as pd
import numpy as np
import openpyxl
import os
import datetime


class Salakabatt:
    """
        - Read the individual excel file
    """
    def __init__(self, file_path, *args, **kwargs):
        self.__file_path = file_path


    def __dataframe(self, data_type: str = 'tax'):
        try:
            wb = openpyxl.load_workbook(self.__file_path, data_only=True, read_only=True)
            tax_sheet = wb.worksheets[0]
            non_tax_sheet = wb.worksheets[1]

            if data_type == 'tax':
                data = []
                for i, row in enumerate(tax_sheet.iter_rows(values_only=True), start=1):
                    data.append(row)
            else:
                """
                data_type == 'non_tax' or any other value will return non-tax data
                """
                data = []
                for i, row in enumerate(non_tax_sheet.iter_rows(values_only=True), start=1):
                    data.append(row)

        except Exception as e:
            pass

        tax_df = pd.DataFrame(data).dropna()
        non_tax_df = pd.DataFrame(data).dropna()

        if data_type == 'tax':
            return tax_df
        return non_tax_df


    def tax_data(self):
        return self.__dataframe(data_type='tax')
    

    def non_tax_data(self):
        return self.__dataframe(data_type='non_tax')


class AllSalakabatt:
    """
        - Read all the excel files
    """
    def __init__(self, list_of_file_path: list, *args, **kwargs):
        self.files = list_of_file_path
        self.tax_data = self.__all_tax_data()
        self.non_tax_data = self.__all_non_tax_data()
        # self.tax_data.apply(pd.to_datetime, errors='ignore')
        # self.non_tax_data.apply(pd.to_datetime, errors='ignore')

    def __all_tax_data(self):
        tax_data_list = []
        for file in self.files:
            tax_data = Salakabatt(file).tax_data()
            tax_data_list.append(tax_data)
        
        combined_tax_data = pd.concat(tax_data_list, ignore_index=True)
        
        return combined_tax_data

    def __all_non_tax_data(self):
        non_tax_data_list = []
        for file in self.files:
            non_tax_data = Salakabatt(file).non_tax_data()
            non_tax_data_list.append(non_tax_data)
        combined_non_tax_data = pd.concat(non_tax_data_list, ignore_index=True)
        
        return combined_non_tax_data

    def income_by_date(self, date_val: datetime.date, income_code: str):
        """
             - Return income on a specific date
        """
        TAX_INCOME_CODE_MAP = {
            'VPP': 4, # 70025
            'VOP': 5, # 70026
            'SPP': 6, # 70032
            'SOP': 7, # 70033
            'COP': 8, # 71001

            'CPP': 10, # 71003
            'ATP': 11, # 71004
            'AUC': 12, # 71005
            'PIM': 13, # 71006
            'ETW': 14, # 71011
            'ETR': 15, # 71012
            'ETP': 16, # 71013
            'ETO': 17, # 71014
            'PEX': 18, # 71016
        }
        NON_TAX_INCOME_CODE_MAP = {
            'SOS': 4, # 73011
            '73012': 5, # 73012
            'SOM': 6, # 73013
            'SOV': 7, # 73015
            'SOD': 8, # 73016
            'SOE': 9, # 73017
            'SOO': 10, # 73018
            'CBF': 11, # 73024
            '73028': 12, # 73028
            '73048': 13, # 73048
            '73066': 14, # 73066
            'ROL': 15, # 73071
            'ROB': 16, # 73072
            'ROO': 17, # 73073
            '73087': 18, # 73087
            'ORB': 19, # 76981
        }

        combined_income_code = {**TAX_INCOME_CODE_MAP, **NON_TAX_INCOME_CODE_MAP}

        if income_code.upper() not in combined_income_code:
            return 0
        if not isinstance(date_val, datetime.datetime):
            raise ValueError(f'Invalid date: {date_val}')

        # Tax Data
        if income_code.upper() in TAX_INCOME_CODE_MAP:
            COLUMN_INDEX = TAX_INCOME_CODE_MAP.get(income_code.upper())
            data = self.tax_data
        # Non-Tax Data
        if income_code.upper() in NON_TAX_INCOME_CODE_MAP:
            COLUMN_INDEX = NON_TAX_INCOME_CODE_MAP.get(income_code.upper())
            data = self.non_tax_data
        
        data.iloc[:, 2] = data.iloc[:, 2].apply(lambda x: datetime.datetime.strptime(x, '%d/%m/%Y') if isinstance(x, str) else x)

        income_by_date = data[data.iloc[:, 2] == date_val][COLUMN_INDEX].sum()
            
        return income_by_date
    

file_path = "/mnt/c/Users/Asus/OneDrive/Desktop/ProjectTest/KAM12_MonthlyReport_Jan_2024.xlsx"

# test = AllSalakabatt([file_path])
# dara = test.income_by_date(datetime.datetime(2024, 4, 1), 'VOP')
# print(dara)
