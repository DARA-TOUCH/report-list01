import os
import shutil
import sys
import openpyxl

from django.conf import settings

__all__ = ['List01']


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, '../../', 'staticfiles', 'excel', 'xlsx'))
TEMPLATE_PATH = os.path.join(PARENT_DIR, 'List01.xlsx')
BUDGET_PATH = os.path.join(PARENT_DIR, 'Budget.xlsx')
SAD_DETAIL_PATH = os.path.join(PARENT_DIR, 'SAD_Detail.xlsx')

MODULE_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, '../'))

sys.path.insert(0, MODULE_DIR)

from kam12filereader.main_operation import BudgetAccount, SADDetail
from kam12list01.salakabatt import AllSalakabatt

class List01:
    def __init__(self, template_file, budget_file, sad_detail_file, list_of_salakabatt_path: list, *args, **kwargs):
        self.template_file = template_file
        self.budget_file = budget_file
        self.sad_detail_file = sad_detail_file
        self.list_of_salakabatt_path = list_of_salakabatt_path

    def write(self, reporter_name: str, office_choice: str):
        list01_wb = openpyxl.load_workbook(self.template_file, data_only=False)
        ws = list01_wb.active
        budget = BudgetAccount(self.budget_file)
        sad_detail = SADDetail(self.sad_detail_file)
        salakabatt = AllSalakabatt(list_of_file_path=self.list_of_salakabatt_path)        
        
        # Temporarily set data_only=True to evaluate all formulas once
        temp_wb = openpyxl.load_workbook(self.template_file, data_only=True)
        # calculated_values = {}

        for sh in temp_wb.sheetnames:
            # calculated_values[sh] = {}
            for row in range(93, 2704, 87):
                list01_wb[sh][f"A{row}"].value = temp_wb[sh][f"A{row}"].value

        SKIP_CELL = ['TRF', 'CPF', 'VAP', '', None] # Cells that it column E contain formulaor blank  should be skipped for calculation

        
        for sh in list01_wb.sheetnames:
            list01_wb[sh]['U3051'].value = reporter_name
            list01_wb[sh]['A1'].value = f'{list01_wb[sh]['A1'].value} {office_choice}'
            for row in range(93, 2704, 87):
                for i in range(1, 87):
                    if list01_wb[sh][f'E{row+i}'].value not in SKIP_CELL:
                        # Revenue (In)
                        list01_wb[sh][f"K{row+i}"].value = budget.get_return_amount_by_date(
                                        income_code=list01_wb[sh][f"E{row+i}"].value,
                                        in_or_out="I",
                                        date_val=list01_wb[sh][f"A{row}"].value
                                        )
                        # Paid to National
                        list01_wb[sh][f"M{row+i}"].value = salakabatt.income_by_date(
                                        income_code=list01_wb[sh][f"E{row+i}"].value,
                                        date_val=list01_wb[sh][f"A{row}"].value
                                        )
        list01_wb.save(self.template_file)